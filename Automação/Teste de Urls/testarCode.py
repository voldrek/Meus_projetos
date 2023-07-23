import openpyxl
import aiohttp
import asyncio
from openpyxl.chart import BarChart, Reference

# Função para verificar o status da URL e retornar o resultado do teste
async def verificar_url(session, url):
    try:
        async with session.get(url) as response:
            if response.status == 200:
                return 'OK'
            else:
                return 'Erro'
    except aiohttp.ClientError:
        return 'Erro'

async def main():
    async with aiohttp.ClientSession() as session:
        # Abrir o arquivo "url.txt" em modo de leitura
        with open('url.txt', 'r') as arquivo:
            # Criar um novo arquivo XLSX
            nome_arquivo_xlsx = "respostas.xlsx"
            workbook = openpyxl.Workbook()
            worksheet = workbook.active

            # Adicionar os cabeçalhos ao arquivo XLSX
            cabecalhos = ["URL", "Status"]
            worksheet.append(cabecalhos)

            # Lista para armazenar os resultados dos testes
            resultados = []

            # Criar uma lista de tarefas assíncronas para processar as URLs em paralelo
            tasks = []
            for url in arquivo:
                url = url.strip()  # Remover espaços em branco no início e no fim da URL
                tasks.append(verificar_url(session, url))

            # Executar as tarefas assíncronas e aguardar os resultados
            resultados = await asyncio.gather(*tasks)

            # Iterar sobre as URLs e gravar os resultados no arquivo XLSX
            arquivo.seek(0)  # Reiniciar o ponteiro do arquivo para ler novamente
            for url, status in zip(arquivo, resultados):
                url = url.strip()  # Remover espaços em branco no início e no fim da URL
                linha = [url, status]
                worksheet.append(linha)

            # Contar o total de URLs testadas, "OK" e "Erro"
            total_ok = resultados.count("OK")
            total_erro = resultados.count("Erro")
            total_testados = total_ok + total_erro

            # Adicionar as colunas com o total de "OK" e "Erro"
            worksheet.insert_cols(3)
            worksheet.cell(row=1, column=3, value="Total OK")
            worksheet.cell(row=2, column=3, value=total_ok)
            worksheet.cell(row=1, column=4, value="Total Erro")
            worksheet.cell(row=2, column=4, value=total_erro)

            # Adicionar a coluna com o total de URLs testadas
            worksheet.insert_cols(5)
            worksheet.cell(row=1, column=5, value="Total URLs")
            worksheet.cell(row=2, column=5, value=total_testados)

            # Criar os dados para o gráfico de barras
            chart = BarChart()
            data = Reference(worksheet, min_col=3, min_row=1, max_row=2, max_col=4)
            categories = Reference(worksheet, min_col=3, min_row=1, max_row=1, max_col=4)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)

            # Definir os títulos dos eixos X e Y
            chart.x_axis.title = 'Status'
            chart.y_axis.title = "Quantidade"

            # Adicionar o gráfico à planilha
            worksheet.add_chart(chart, "G1")

            # Salvar o arquivo XLSX com o gráfico de barras
            workbook.save(nome_arquivo_xlsx)

        print(f"Os resultados e o gráfico foram gravados em '{nome_arquivo_xlsx}' com sucesso!")

if __name__ == "__main__":
    asyncio.run(main())
